#!/usr/bin/env python3
"""
Script to process Excel files - reads input Excel, adds a processed column, and saves to output.
Demonstrates working with Excel files using openpyxl.
"""
import argparse
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    print("ERROR: 'openpyxl' library not installed")
    print("Please install it with: pip install openpyxl")
    sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description='Process Excel file - add processed column')
    parser.add_argument('--input', type=str, required=True, help='Input Excel file path')
    parser.add_argument('--output', type=str, default='output.xlsx', help='Output Excel file path')

    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    print(f"📊 Processing Excel file: {input_path}")
    print("-" * 70)

    if not input_path.exists():
        print(f"❌ ERROR: Input file not found: {input_path}")
        sys.exit(1)

    try:
        # Load workbook
        print(f"📖 Loading workbook...")
        wb = load_workbook(input_path)

        # Get active sheet
        ws = wb.active
        sheet_name = ws.title
        print(f"✅ Loaded sheet: '{sheet_name}'")

        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"📐 Dimensions: {max_row} rows × {max_col} columns")

        # Print column names (first row)
        print(f"\n📋 Column names:")
        columns = []
        for col in range(1, max_col + 1):
            cell_value = ws.cell(row=1, column=col).value
            columns.append(cell_value)
            print(f"  Column {col}: {cell_value}")

        # Add "processed" column
        processed_col = max_col + 1
        ws.cell(row=1, column=processed_col, value="processed")
        print(f"\n✨ Adding 'processed' column at position {processed_col}")

        # Fill "processed" column with True for all data rows
        for row in range(2, max_row + 1):
            ws.cell(row=row, column=processed_col, value=True)

        print(f"✅ Filled {max_row - 1} rows with processed=True")

        # Save to output
        wb.save(output_path)
        print(f"\n💾 Saved output to: {output_path.absolute()}")

        print("-" * 70)
        print(f"✅ Excel processing completed successfully!")
        print(f"   Input:  {input_path}")
        print(f"   Output: {output_path}")
        sys.exit(0)

    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
